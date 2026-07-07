"""
Student Notebook & Project Submission Tracker
Flask backend. Uses SQLite for local development by default, or a real
Postgres database (e.g. a free Neon instance) when a DATABASE_URL
environment variable is set -- so data isn't tied to Render's ephemeral
disk / plan at all, and persists until you explicitly clear it.
"""
import os
import re
import io
import csv
import sqlite3
from datetime import datetime

from flask import Flask, request, jsonify, render_template, send_file, g
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
    IntegrityError = psycopg2.IntegrityError
else:
    IntegrityError = sqlite3.IntegrityError

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# DATA_DIR can be pointed at a Render persistent disk mount (e.g. /var/data)
# for the SQLite file, and is always used for uploads/exports regardless of
# which database backend is active. Defaults to the project folder locally.
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
DB_PATH = os.path.join(DATA_DIR, "tracker.db")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
EXPORT_DIR = os.path.join(DATA_DIR, "exports")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25 MB

SHEETS_TO_IGNORE = {"summary", "index", "readme", "cover"}


# ---------------------------------------------------------------------------
# Database helpers
#
# A thin wrapper gives the rest of the app one uniform API
# (execute/executescript/commit/close) regardless of whether the backend is
# SQLite (local dev) or Postgres (production, e.g. Neon). "?" placeholders
# are translated to "%s" for Postgres. Rows behave like dicts either way
# (sqlite3.Row / psycopg2 RealDictCursor), so `row["col"]` and `dict(row)`
# both work unchanged everywhere else in this file.
# ---------------------------------------------------------------------------
class DBWrapper:
    def __init__(self, conn):
        self._conn = conn

    def execute(self, sql, params=()):
        if USE_POSTGRES:
            sql = sql.replace("?", "%s")
            cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(sql, params)
            return cur
        return self._conn.execute(sql, params)

    def executescript(self, script):
        if USE_POSTGRES:
            cur = self._conn.cursor()
            cur.execute(script)
        else:
            self._conn.executescript(script)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()


def _raw_connect():
    if USE_POSTGRES:
        return psycopg2.connect(DATABASE_URL)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def get_db():
    if "db" not in g:
        g.db = DBWrapper(_raw_connect())
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    conn = DBWrapper(_raw_connect())
    if USE_POSTGRES:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS students (
                id SERIAL PRIMARY KEY,
                class_name TEXT NOT NULL,
                section TEXT NOT NULL,
                roll_no INTEGER,
                name TEXT NOT NULL,
                notebook INTEGER DEFAULT 0,
                project INTEGER DEFAULT 0,
                updated_at TEXT,
                UNIQUE(class_name, section, roll_no, name)
            );

            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY,
                value TEXT
            );
            """
        )
    else:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_name TEXT NOT NULL,
                section TEXT NOT NULL,
                roll_no INTEGER,
                name TEXT NOT NULL,
                notebook INTEGER DEFAULT 0,
                project INTEGER DEFAULT 0,
                updated_at TEXT,
                UNIQUE(class_name, section, roll_no, name)
            );

            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY,
                value TEXT
            );
            """
        )
    conn.commit()
    conn.close()


def set_meta(key, value):
    db = get_db()
    db.execute(
        "INSERT INTO meta (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )
    db.commit()


def get_meta(key, default=None):
    db = get_db()
    row = db.execute("SELECT value FROM meta WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


# ---------------------------------------------------------------------------
# Excel import parsing
# ---------------------------------------------------------------------------
def parse_sheet_name(sheet_name):
    """
    Parses a sheet name like 'VI - Ravi', 'VI Ravi', 'VII-Alaknanda'
    into (class_name, section).
    """
    name = sheet_name.strip()
    # Try 'Class - Section' or 'Class-Section'
    m = re.match(r"^\s*([A-Za-z]+)\s*-\s*(.+?)\s*$", name)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    # Try 'Class Section' (space separated, class is first token)
    parts = name.split(None, 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return name, "General"


def find_header_row(rows):
    """
    Given a list of row-tuples for a sheet, find the index (0-based) of the
    row that looks like a header (contains 'S.No' / 'Roll' and 'Name').
    Returns the index of the first DATA row (header_row_index + 1),
    or None if no usable header/data found.
    """
    for i, row in enumerate(rows[:8]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)
        if ("s.no" in joined or "roll" in joined or "sno" in joined) and "name" in joined:
            return i + 1
    return None


def import_workbook(filepath):
    """Reads every sheet, extracts class/section/roll/name, saves to DB."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    db = get_db()

    sheets_imported = 0
    students_added = 0
    students_skipped_dupe = 0
    sections_seen = set()

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in SHEETS_TO_IGNORE:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        data_start = find_header_row(rows)
        if data_start is None:
            # No recognizable header — assume row 0 is header, data from row 1
            data_start = 1

        class_name, section = parse_sheet_name(sheet_name)
        sections_seen.add((class_name, section))
        sheets_imported += 1

        auto_roll = 0
        for row in rows[data_start:]:
            if row is None:
                continue
            cells = list(row) + [None, None]
            roll_raw, name_raw = cells[0], cells[1]

            # Skip fully blank rows
            if (roll_raw is None or str(roll_raw).strip() == "") and (
                name_raw is None or str(name_raw).strip() == ""
            ):
                continue

            name_val = str(name_raw).strip() if name_raw is not None else ""
            # Skip placeholder / notice rows (no real student name, e.g.
            # "No student list available for this section...")
            if not name_val or (roll_raw is None and len(name_val.split()) > 6):
                continue
            if roll_raw is None and any(
                kw in name_val.lower()
                for kw in ["no student", "not available", "missing", "photo"]
            ):
                continue

            try:
                roll_no = int(roll_raw)
            except (TypeError, ValueError):
                auto_roll += 1
                roll_no = auto_roll
            else:
                auto_roll = roll_no

            try:
                row = db.execute(
                    "INSERT INTO students (class_name, section, roll_no, name, "
                    "notebook, project, updated_at) VALUES (?, ?, ?, ?, 0, 0, ?) "
                    "ON CONFLICT (class_name, section, roll_no, name) DO NOTHING "
                    "RETURNING id",
                    (class_name, section, roll_no, name_val, datetime.now().isoformat()),
                ).fetchone()
            except IntegrityError:
                db.rollback()
                row = None

            if row is not None:
                students_added += 1
            else:
                students_skipped_dupe += 1

    db.commit()

    set_meta("last_import_date", datetime.now().strftime("%Y-%m-%d %H:%M"))

    return {
        "sheets_imported": sheets_imported,
        "students_added": students_added,
        "duplicates_skipped": students_skipped_dupe,
        "sections": sorted(sections_seen),
    }


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------
@app.route("/")
def home():
    return render_template("index.html")


@app.route("/dashboard")
def dashboard_page():
    return render_template("dashboard.html")


@app.route("/students-page")
def students_page():
    return render_template("students.html")


# ---------------------------------------------------------------------------
# Upload / Import
# ---------------------------------------------------------------------------
@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"success": False, "message": "No file part in request"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "message": "No file selected"}), 400
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"success": False, "message": "Please upload a .xlsx file"}), 400

    save_path = os.path.join(UPLOAD_DIR, "Students.xlsx")
    file.save(save_path)
    return jsonify({"success": True, "message": "File uploaded. Now importing...", "path": save_path})


@app.route("/import", methods=["POST"])
def do_import():
    replace = request.args.get("replace", "false").lower() == "true" or (
        bool((request.get_json(silent=True) or {}).get("replace")) if request.is_json else False
    )
    filepath = os.path.join(UPLOAD_DIR, "Students.xlsx")
    if not os.path.exists(filepath):
        return jsonify({"success": False, "message": "No workbook found. Please upload first."}), 400

    db = get_db()
    if replace:
        db.execute("DELETE FROM students")
        db.commit()

    try:
        result = import_workbook(filepath)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"success": False, "message": f"Import failed: {exc}"}), 500

    total_students = db.execute("SELECT COUNT(*) c FROM students").fetchone()["c"]
    set_meta("total_students", str(total_students))

    summary = (
        f"{result['sheets_imported']} sheets imported, "
        f"{result['students_added']} students added"
    )
    if result["duplicates_skipped"]:
        summary += f" ({result['duplicates_skipped']} duplicates skipped)"

    return jsonify(
        {
            "success": True,
            "message": summary,
            "sheets_imported": result["sheets_imported"],
            "students_added": result["students_added"],
            "duplicates_skipped": result["duplicates_skipped"],
            "total_students": total_students,
        }
    )


@app.route("/api/settings", methods=["GET"])
def get_settings():
    return jsonify(
        {
            "school_name": get_meta("school_name", "My School"),
            "academic_session": get_meta("academic_session", "2025-26"),
        }
    )


@app.route("/api/settings", methods=["POST"])
def save_settings():
    data = request.get_json(force=True)
    school_name = (data.get("school_name") or "").strip()
    academic_session = (data.get("academic_session") or "").strip()

    if school_name:
        set_meta("school_name", school_name)
    if academic_session:
        set_meta("academic_session", academic_session)

    return jsonify(
        {
            "success": True,
            "school_name": get_meta("school_name", "My School"),
            "academic_session": get_meta("academic_session", "2025-26"),
        }
    )


# ---------------------------------------------------------------------------
# Class / Section / Student data APIs
# ---------------------------------------------------------------------------
@app.route("/api/classes")
def api_classes():
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT class_name FROM students ORDER BY class_name"
    ).fetchall()
    classes = [r["class_name"] for r in rows]
    # Sort numerals/roman-numeral-ish classes sensibly
    roman_order = ["VI", "VII", "VIII", "IX", "X", "XI", "XII"]
    classes.sort(key=lambda c: roman_order.index(c) if c in roman_order else 99)
    return jsonify(classes)


@app.route("/api/sections")
def api_sections():
    class_name = request.args.get("class_name", "")
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT section FROM students WHERE class_name=? ORDER BY section",
        (class_name,),
    ).fetchall()
    return jsonify([r["section"] for r in rows])


@app.route("/students")
def get_students():
    class_name = request.args.get("class_name", "")
    section = request.args.get("section", "")
    search = request.args.get("search", "").strip()
    roll_search = request.args.get("roll_search", "").strip()
    filter_type = request.args.get("filter", "all")

    query = "SELECT * FROM students WHERE class_name=? AND section=?"
    params = [class_name, section]

    if search:
        query += " AND name LIKE ?"
        params.append(f"%{search}%")
    if roll_search:
        query += " AND CAST(roll_no AS TEXT) LIKE ?"
        params.append(f"%{roll_search}%")

    if filter_type == "pending_notebook":
        query += " AND notebook=0"
    elif filter_type == "pending_project":
        query += " AND project=0"
    elif filter_type == "completed":
        query += " AND notebook=1 AND project=1"
    elif filter_type == "pending":
        query += " AND (notebook=0 OR project=0)"

    query += " ORDER BY roll_no"

    db = get_db()
    rows = db.execute(query, params).fetchall()
    students = [dict(r) for r in rows]

    total = len(students)
    notebook_done = sum(1 for s in students if s["notebook"])
    project_done = sum(1 for s in students if s["project"])

    return jsonify(
        {
            "students": students,
            "summary": {
                "total": total,
                "notebook_submitted": notebook_done,
                "notebook_pending": total - notebook_done,
                "project_submitted": project_done,
                "project_pending": total - project_done,
                "notebook_pct": round((notebook_done / total * 100) if total else 0, 1),
                "project_pct": round((project_done / total * 100) if total else 0, 1),
            },
        }
    )


@app.route("/update_status", methods=["POST"])
def update_status():
    data = request.get_json(force=True)
    student_id = data.get("id")
    field = data.get("field")  # 'notebook' or 'project'
    value = 1 if data.get("value") else 0

    if field not in ("notebook", "project"):
        return jsonify({"success": False, "message": "Invalid field"}), 400

    db = get_db()
    db.execute(
        f"UPDATE students SET {field}=?, updated_at=? WHERE id=?",
        (value, datetime.now().isoformat(), student_id),
    )
    db.commit()
    row = db.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    return jsonify({"success": True, "student": dict(row) if row else None})


@app.route("/bulk_action", methods=["POST"])
def bulk_action():
    data = request.get_json(force=True)
    action = data.get("action")
    class_name = data.get("class_name")
    section = data.get("section")

    db = get_db()
    now = datetime.now().isoformat()

    actions = {
        "mark_all_notebook": ("UPDATE students SET notebook=1, updated_at=? WHERE class_name=? AND section=?", (now, class_name, section)),
        "mark_all_project": ("UPDATE students SET project=1, updated_at=? WHERE class_name=? AND section=?", (now, class_name, section)),
        "clear_notebook": ("UPDATE students SET notebook=0, updated_at=? WHERE class_name=? AND section=?", (now, class_name, section)),
        "clear_project": ("UPDATE students SET project=0, updated_at=? WHERE class_name=? AND section=?", (now, class_name, section)),
        "reset_section": ("UPDATE students SET notebook=0, project=0, updated_at=? WHERE class_name=? AND section=?", (now, class_name, section)),
    }

    if action not in actions:
        return jsonify({"success": False, "message": "Unknown action"}), 400

    sql, params = actions[action]
    db.execute(sql, params)
    db.commit()
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Dashboard summary
# ---------------------------------------------------------------------------
@app.route("/summary")
def summary():
    db = get_db()

    overall = db.execute(
        """SELECT COUNT(*) total,
                  SUM(notebook) notebook_submitted,
                  SUM(project) project_submitted
           FROM students"""
    ).fetchone()
    total = overall["total"] or 0
    notebook_submitted = overall["notebook_submitted"] or 0
    project_submitted = overall["project_submitted"] or 0

    class_rows = db.execute(
        """SELECT class_name,
                  COUNT(*) students,
                  SUM(notebook) notebook_done,
                  SUM(project) project_done
           FROM students GROUP BY class_name ORDER BY class_name"""
    ).fetchall()

    section_rows = db.execute(
        """SELECT class_name, section,
                  COUNT(*) students,
                  SUM(notebook) notebook_done,
                  SUM(project) project_done
           FROM students GROUP BY class_name, section
           ORDER BY class_name, section"""
    ).fetchall()

    def pct(done, tot):
        return round((done or 0) / tot * 100, 1) if tot else 0.0

    class_summary = [
        {
            "class_name": r["class_name"],
            "students": r["students"],
            "notebook_pct": pct(r["notebook_done"], r["students"]),
            "project_pct": pct(r["project_done"], r["students"]),
        }
        for r in class_rows
    ]

    section_summary = [
        {
            "class_name": r["class_name"],
            "section": r["section"],
            "students": r["students"],
            "notebook_done": r["notebook_done"] or 0,
            "project_done": r["project_done"] or 0,
            "pending": r["students"] - min(r["notebook_done"] or 0, r["project_done"] or 0),
        }
        for r in section_rows
    ]

    return jsonify(
        {
            "school_name": get_meta("school_name", "My School"),
            "academic_session": get_meta("academic_session", "2025-26"),
            "last_import_date": get_meta("last_import_date", "Never"),
            "overall": {
                "total_students": total,
                "notebook_submitted": notebook_submitted,
                "notebook_pending": total - notebook_submitted,
                "project_submitted": project_submitted,
                "project_pending": total - project_submitted,
            },
            "class_summary": class_summary,
            "section_summary": section_summary,
        }
    )


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------
@app.route("/export/excel")
def export_excel():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM students ORDER BY class_name, section, roll_no"
    ).fetchall()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sections = {}
    for r in rows:
        key = f"{r['class_name']} - {r['section']}"
        sections.setdefault(key, []).append(r)

    for sheet_name, students in sections.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(["S.No.", "Student Name", "Notebook Submitted", "Project Submitted", "Status"])
        for s in students:
            status = (
                "Complete" if s["notebook"] and s["project"]
                else "Partial" if (s["notebook"] or s["project"])
                else "Pending"
            )
            ws.append([
                s["roll_no"], s["name"],
                "Yes" if s["notebook"] else "No",
                "Yes" if s["project"] else "No",
                status,
            ])
        for col_idx, width in enumerate([8, 30, 20, 20, 12], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    out_path = os.path.join(EXPORT_DIR, "Submission_Report.xlsx")
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name="Submission_Report.xlsx")


@app.route("/export/csv")
def export_csv():
    class_name = request.args.get("class_name")
    section = request.args.get("section")

    db = get_db()
    query = "SELECT * FROM students"
    params = []
    conditions = []
    if class_name:
        conditions.append("class_name=?")
        params.append(class_name)
    if section:
        conditions.append("section=?")
        params.append(section)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY class_name, section, roll_no"

    rows = db.execute(query, params).fetchall()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Class", "Section", "Roll No", "Name", "Notebook", "Project", "Status"])
    for r in rows:
        status = (
            "Complete" if r["notebook"] and r["project"]
            else "Partial" if (r["notebook"] or r["project"])
            else "Pending"
        )
        writer.writerow([
            r["class_name"], r["section"], r["roll_no"], r["name"],
            "Yes" if r["notebook"] else "No",
            "Yes" if r["project"] else "No",
            status,
        ])

    mem = io.BytesIO(buf.getvalue().encode("utf-8"))
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name="Submission_Report.csv")


@app.route("/export/pdf")
def export_pdf():
    scope = request.args.get("scope", "school")  # school | class | section
    class_name = request.args.get("class_name")
    section = request.args.get("section")

    db = get_db()
    query = "SELECT * FROM students"
    params = []
    conditions = []
    title = "Overall School Report"
    if scope == "class" and class_name:
        conditions.append("class_name=?")
        params.append(class_name)
        title = f"Class {class_name} Report"
    elif scope == "section" and class_name and section:
        conditions.append("class_name=?")
        conditions.append("section=?")
        params.extend([class_name, section])
        title = f"Class {class_name} - Section {section} Report"
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY class_name, section, roll_no"

    rows = db.execute(query, params).fetchall()

    out_path = os.path.join(EXPORT_DIR, "Submission_Report.pdf")
    doc = SimpleDocTemplate(out_path, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 0.5 * cm)]

    data = [["Class", "Section", "Roll No", "Name", "Notebook", "Project", "Status"]]
    for r in rows:
        status = (
            "Complete" if r["notebook"] and r["project"]
            else "Partial" if (r["notebook"] or r["project"])
            else "Pending"
        )
        data.append([
            r["class_name"], r["section"], str(r["roll_no"]), r["name"],
            "Yes" if r["notebook"] else "No",
            "Yes" if r["project"] else "No",
            status,
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2FB")]),
    ]))
    elements.append(table)
    doc.build(elements)

    return send_file(out_path, as_attachment=True, download_name=f"{title.replace(' ', '_')}.pdf")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.environ.get("FLASK_DEBUG", "true").lower() == "true"
    # use_reloader=False: the reloader watches the whole project folder,
    # and saving uploaded/exported files inside it (uploads/, exports/,
    # tracker.db) would otherwise trigger a reload mid-request and kill
    # the connection (seen as random 500s on /upload).
    app.run(debug=debug_mode, host="0.0.0.0", port=port, use_reloader=False)
else:
    init_db()
